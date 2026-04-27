import openpyxl
import datetime
import os
from spire.xls import *
from spire.xls.common import *
import openpyxl
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog



class Itm:
	def __init__(self,name,price):
		self.name=name
		self.price=price

class CurrentItem:
	def __init__(self,name,price,amount):
		self.name=name
		self.price=price
		self.amount=amount








#global variables
version = "v1.0"
title = ""
name = ""
address = ""
phone = ""
date = datetime.date.today().strftime('%Y.%m.%d')
height = 400
width = 720
num = 0
mergeValue = False
excelValue = True
pdfValue = False
picValue = False
excelCheckG = 1
pdfCheckG = 0
picCheckG = 0
currentItems = ".\items\default.itm"
itemList = []
setCurrent = CurrentItem('0',0,0)

#golbal functions
def manageClick():
	mamageWindow = tk.Tk()
	mamageWindow.geometry("480x360")
	mamageWindow.resizable(False, False)
	mamageWindow.title("类目物品管理")
	mamageWindow.iconbitmap("icon.ico")

def importExportClick():
	global mainWindow
	mainWindow.withdraw()
	ieWindow = tk.Toplevel(mainWindow)
	ieWindow.lift()
	ieWindow.protocol("WM_DELETE_WINDOW",lambda:disp(ieWindow))
	ieWindow.geometry("480x260")
	ieWindow.resizable(False, False)
	ieWindow.title("打开、保存和导出")
	ieWindow.iconbitmap("icon.ico")
	btnSave = tk.Button(ieWindow, text="保存", width=int(width/50)-2, height=1, command=lambda:saveFile(entSave), compound="c")
	btnImport = tk.Button(ieWindow, text="导入", width=int(width/50)-2, height=1, command=lambda:importItem(entImport), compound="c")
	btnExport = tk.Button(ieWindow, text="导出", width=int(width/50)-2, height=1, command=lambda:exportFile(entExport), compound="c")
	btnBrowseSave = tk.Button(ieWindow, text="浏览", width=int(width/50)-2, height=1, command=lambda:browseSave(entSave), compound="c")
	btnBrowseImport = tk.Button(ieWindow, text="浏览", width=int(width/50)-2, height=1, command=lambda:browseImport(entImport), compound="c")
	btnBrowseExport = tk.Button(ieWindow, text="浏览", width=int(width/50)-2, height=1, command=lambda:browseExport(entExport), compound="c")
	lblSave = tk.Label(ieWindow,text='保存当前编工程到：')
	lblImport = tk.Label(ieWindow,text='导入工程文件：')
	lblExport = tk.Label(ieWindow,text='导出到：')
	btnCancel = tk.Button(ieWindow, text="取消", width=int(width/50)-2, height=1, command=lambda:cancel(ieWindow), compound="c")
	entSave = tk.Entry(ieWindow, width=int(width/18))
	entImport = tk.Entry(ieWindow, width=int(width/18))
	entExport = tk.Entry(ieWindow, width=int(width/18))
	excelCheck = tk.IntVar()
	excelCheck.set(1)
	pdfCheck = tk.IntVar()
	picCheck = tk.IntVar()
	ckbExcel = tk.Checkbutton(ieWindow, text="导出Excel", variable=excelCheck, onvalue=1, offvalue=0, command=exExcel, height=1)
	ckbPDF = tk.Checkbutton(ieWindow, text="导出PDF", variable=pdfCheck, onvalue=1, offvalue=0, command=exPDF, height=1)
	ckbPic = tk.Checkbutton(ieWindow, text="导出图片", variable=picCheck, onvalue=1, offvalue=0, command=exPic, height=1)
	lblSave.place(x=0, y=0*int(height/13)+10, anchor='nw')
	lblImport.place(x=0, y=2*int(height/13)+10, anchor='nw')
	lblExport.place(x=0, y=4*int(height/13)+10, anchor='nw')
	entSave.place(x=0, y=1*int(height/13)+10, anchor='nw')
	entImport.place(x=0, y=3*int(height/13)+10, anchor='nw')
	entExport.place(x=0, y=5*int(height/13)+10, anchor='nw')
	btnBrowseSave.place(x=285, y=1*int(height/13)+5, anchor='nw')
	btnBrowseImport.place(x=285, y=3*int(height/13)+5, anchor='nw')
	btnBrowseExport.place(x=285, y=5*int(height/13)+5, anchor='nw')
	btnSave.place(x=380, y=1*int(height/13)+5, anchor='nw')
	btnImport.place(x=380, y=3*int(height/13)+5, anchor='nw')
	btnExport.place(x=380, y=5*int(height/13)+5, anchor='nw')
	ckbExcel.place(x=10, y=7*int(height/13)+10, anchor='nw')	
	ckbPDF.place(x=130, y=7*int(height/13)+10, anchor='nw')	
	ckbPic.place(x=250, y=7*int(height/13)+10, anchor='nw')	
	btnCancel.place(x=380, y=7*int(height/13)+10, anchor='nw')
	ieWindow.mainloop()

def disp(wind):
	global mainWindow
	wind.destroy()
	mainWindow.deiconify()

def exExcel():
	global excelCheckG
	global excelValue
	if excelCheckG == 0:
		excelCheckG = 1
	else:
		excelCheckG = 0
	print(excelCheckG)
	global excelValue
	if excelCheckG == 1:
		excelValue = True
	else:
		excelValue = False
	print(excelCheckG)
	print(excelValue)

def exPDF():
	global pdfCheckG
	global pdfValue
	if pdfCheckG == 0:
		pdfCheckG = 1
	else:
		pdfCheckG = 0
	if pdfCheckG == 1:
		pdfValue = True
	else:
		pdfValue = False
	print(pdfCheckG)
	print(pdfValue)

def exPic():
	global picCheckG
	global picValue
	if picCheckG == 0:
		picCheckG = 1
	else:
		picCheckG = 0
	if picCheckG == 1:
		picValue = True
	else:
		picValue = False
	print(picCheckG)
	print(picValue)

def browseSave(entSave):
	global entTitle
	filePath=filedialog.asksaveasfilename(initialdir=".\\"+"saves"+"\\",initialfile=entTitle.get()+"_"+datetime.datetime.today().strftime('%Y_%m_%d_%H_%M_%S')+".qtn",defaultextension=".qtn",filetypes=[("QTN Files","*.qtn"),("All Files","*.*")])
	entSave.delete("0","end")
	entSave.insert("end",filePath)

def browseImport(entImport):
	filePath=filedialog.askopenfilename(defaultextension=".qtn",filetypes=[("QTN Files","*.qtn"),("All Files","*.*")])
	if filePath:
		entImport.delete("0","end")
		entImport.insert("end",filePath)
	else:
		messagebox.showerror(title='错误', message='请选择正确的路径与格式！')

def browseExport(entExport):
	filePath=filedialog.askdirectory()
	if filePath:
		entExport.delete("0","end")
		entExport.insert("end",filePath)
	else:
		messagebox.showerror(title='错误', message='请选择正确的保存路径与合法的文件名！')

def saveFile(entSave):
	global title
	global itemList
	global entTitle
	global entAddress
	global entName
	global entPhone
	global entDate
	global txtContent
	if entSave.get() != "":
		fileTarget = entSave.get()
		fileName = fileTarget.split("\\")[-1]
		print(fileTarget)
		print(fileName)
		if fileName != "":
			fd = os.open(fileTarget,os.O_WRONLY | os.O_CREAT)
			os.close(fd)
			os.remove(fileTarget)
			fd = os.open(fileTarget,os.O_WRONLY | os.O_CREAT)
			content = entTitle.get() + '\n'  + entAddress.get() + '\n'  + entName.get() + '\n'  + entPhone.get() + '\n' + entDate.get() + '\n'  + txtContent.get("1.0","end")
			for i in itemList:
				content += i.name
				content += ' '
				content += str(i.price)
				content += ' '
				content += str(i.amount)
				content += '\n'
			os.write(fd,content.encode("utf-8"))
			os.close(fd)
		else:
			messagebox.showerror(title='错误', message='请选择正确的保存路径与合法的文件名！')
	else:
		fileTarget = ".\\" + "saves" + "\\" + entTitle.get() + "_" + datetime.datetime.today().strftime('%Y_%m_%d_%H_%M_%S') + ".qtn"
		fileName = fileTarget.split("\\")[-1]
		print(fileTarget)
		print(fileName)
		fd = os.open(fileTarget,os.O_WRONLY | os.O_CREAT)
		content = entTitle.get() + '\n' + entAddress.get() + '\n' + entName.get() + '\n' + entPhone.get() + '\n' + entDate.get() + '\n' + txtContent.get("1.0","end")
		for i in itemList:
			content += i.name
			content += ' '
			content += str(i.price)
			content += ' '
			content += str(i.amount)
			content += '\n'
		os.write(fd,content.encode("utf-8"))
		os.close(fd)

def importItem(entImport):
	global title
	global address
	global name
	global phone
	global date
	global content
	global itemList
	global entTitle
	global entAddress
	global entName
	global entPhone
	global entDate
	global txtContent
	if entImport.get() != []:
		fileTarget = entImport.get()
		fileName = fileTarget.split("\\")[-1]
		print(fileTarget)
		print(fileName)
		itemList = []
		if fileName != "":
			fd = open(fileTarget)
			title = fd.readline()
			entTitle.delete("0","end")
			entTitle.insert("end",title)
			entAddress.delete("0","end")
			entAddress.insert("end",fd.readline())
			entName.delete("0","end")
			entName.insert("end",fd.readline())
			entPhone.delete("0","end")
			entPhone.insert("end",fd.readline())
			entDate.delete("0","end")
			entDate.insert("end",fd.readline())
			txtContent.delete("0.0","end")
			txtContent.insert("end",fd.readline())
			line = fd.readline()
			while line:
				curItm = line.split()
				if len(curItm) >= 3:
					current = CurrentItem(curItm[0],int(curItm[1]),int(curItm[2]))
					itemList.append(current)
				line = fd.readline()
			fd.close()
			refresh()
		else:
			messagebox.showerror(title='错误', message='请选择正确的保存路径与合法的文件名！')

def exportFile(entExport):
	global excelValue
	global pdfValue
	global picValue
	global entTitle
	global entAddress
	global entName
	global entPhone
	global entDate
	global txtContent
	global itemList
	fName = entTitle.get().strip() + "_" + datetime.datetime.today().strftime('%Y_%m_%d_%H_%M_%S')
	if entExport.get() == "":
		exPath = ".\\" + "output" + "\\"
	else:
		exPath = entExport.get() + "\\"
	if excelValue == False and pdfValue == False and picValue == False:
		messagebox.showerror(title='错误', message='请选择导出文件类型！')
	else:
		os.system("copy quotation.xlsx .\\tmp\\quotation.xlsx")
		workbook = load_workbook(filename=".\\tmp\\quotation.xlsx")
		sheet = workbook["Sheet1"]
		print(sheet["A1"].value)
		sheet["A1"].value = entTitle.get()
		sheet["B2"].value = entAddress.get()
		sheet["F2"].value = entPhone.get()
		sheet["I2"].value = entName.get()
		sheet["G49"].value = entDate.get()
		sheet["B45"].value = txtContent.get("0.0","end")
		if len(itemList) <= 40:
			row = 4
			for i in itemList:
				sheet['B'+str(row)] = i.name
				sheet['C'+str(row)] = i.price
				sheet['D'+str(row)] = i.amount
				sheet['E'+str(row)] = i.price * i.amount
				row += 1
		else:
			row = 4
			for i in range(40):
				sheet['B'+str(row)] = itemList[i].name
				sheet['C'+str(row)] = itemList[i].price
				sheet['D'+str(row)] = itemList[i].amount
				sheet['E'+str(row)] = itemList[i].price * itemList[i].amount
				row += 1
			row = 4
			for i in range(40,len(itemList)):
				sheet['G'+str(row)] = itemList[i].name
				sheet['H'+str(row)] = itemList[i].price
				sheet['I'+str(row)] = itemList[i].amount
				sheet['J'+str(row)] = itemList[i].price * itemList[i].amount
				row += 1
		workbook.save(".\\tmp\\quotation1.xlsx")
		workbook.close()
		if excelValue:
			os.system("copy .\\tmp\\quotation1.xlsx " + exPath + fName + ".xlsx")
		if pdfValue:
			wb = Workbook()
			wb.LoadFromFile(".\\tmp\\quotation1.xlsx")
			wb.SaveToFile(exPath + fName + ".pdf", FileFormat.PDF)
			wb.Dispose()
		if picValue:
			print("pic")
		os.remove(".\\tmp\\quotation1.xlsx")
		os.remove(".\\tmp\\quotation.xlsx")

def saveItem():
	print("save")

def openClass():
	filePath = ""
	options = {
		'title': openFileTitle,
		'initialdir': '/',
		'filetypes': [('Supported Files', '*.itm'), ('Item Files', '*.itm'), ('All Files', '*.*')],
		'defaultextension': '*.itm'
	}
	file = tk.filedialog.askopenfiles(mode='r', **options)
	if file is not None and filePath.split('.')[-1] == "itm":
		return file
	else:
		print("请选择正确格式的文件！")
		messagebox.showerror(title='错误', message='请选择正确格式的文件！')
		return None

def openSave():
	filePath = ""
	options = {
		'title': openFileTitle,
		'initialdir': '/',
		'filetypes': [('Supported Files', '*.qtn'), ('Item Files', '*.qtn'), ('All Files', '*.*')],
		'defaultextension': '*.qtn'
	}
	file = tk.filedialog.askopenfiles(mode='r', **options)
	if file is not None and filePath.split('.')[-1] == "qtn":
		return file
	else:
		print("请选择正确格式的文件！")
		messagebox.showerror(title='错误', message='请选择正确格式的文件！')
		return None

def merge():
	global mergeValue
	if mergeCheck.get() == 1:
		mergeValue = True
	else:
		mergeValue = False
	print(mergeCheck)
	print(mergeValue)

def refresh():
	global itemList
	global treeview
	treeview.delete(*treeview.get_children())
	for i in range(len(itemList)):
		treeview.insert("","end",values=(itemList[i].name,str(itemList[i].price),str(itemList[i].amount),str(itemList[i].price*itemList[i].amount)))

def itemInsert():
	global itemList
	global cbbItem
	global entPrice
	global entAmount
	print("insert")
	if mergeValue == True:
		found = False
		for i in itemList:
			if i.name == cbbItem.get() and i.price == int(entPrice.get()):
				i.amount += int(entAmount.get())
				found = True
				break
		if found == False:
			if len(itemList) >= 80:
				messagebox.showerror(title='错误', message='超出长度限制！列表最多可存储80项。')
			else:
				if cbbItem.get() != "" and entPrice.get() != "" and entAmount.get() != "":
					currentItem = CurrentItem(cbbItem.get(),int(entPrice.get()),int(entAmount.get()))
					itemList.append(currentItem)
	else:
		if len(itemList) >= 80:
			messagebox.showerror(title='错误', message='超出长度限制！列表最多可存储80项。')
		else:
			if cbbItem.get() != "" and entPrice.get() != "" and entAmount.get() != "":
				currentItem = CurrentItem(cbbItem.get(),int(entPrice.get()),int(entAmount.get()))
				itemList.append(currentItem)
	refresh()

def itemDelete():
	print("delete")
	global itemList
	selected = treeview.selection()
	print(selected)
	if selected != ():
		selectedValues=treeview.item(selected[0])['values']
		currentItem = CurrentItem(selectedValues[0],int(selectedValues[1]),int(selectedValues[2]))
		print(selectedValues)
		for i in itemList:
			if i.name == currentItem.name and i.price == currentItem.price and i.amount == currentItem.amount:
				itemList.remove(i)
				break
	refresh()

def setItem():
	global setCurrent
	print("set")
	selected = treeview.selection()
	selectedValues=treeview.item(selected[0])['values']
	currentItem = CurrentItem(selectedValues[0],int(selectedValues[1]),int(selectedValues[2]))
	print(selectedValues)
	for i in range(len(itemList)):
		if itemList[i].name == currentItem.name and itemList[i].price == currentItem.price and itemList[i].amount == currentItem.amount:
			itemList.pop(i)
			itemList.insert(i,CurrentItem(cbbCurrentItem.get(),int(entCurrentPrice.get()),int(entCurrentAmount.get())))
			break
	editWindow.destroy()
	refresh()

def cancel(window):
	window.destroy()

def itemEdit():
	global cbbCurrentItem
	global entCurrentPrice
	global entCurrentAmount
	global editWindow
	selected = treeview.selection()
	if selected != ():
		selectedValues=treeview.item(selected[0])['values']
		editWindow = tk.Tk()
		editWindow.geometry("480x240")
		editWindow.resizable(False, False)
		editWindow.title("修改当前项目")
		editWindow.iconbitmap("icon.ico")
		lblCurrentItem = tk.Label(editWindow,text='当前项目：')
		lblCurrentItem2 = tk.Label(editWindow,text=selectedValues[0])
		lblCurrentPrice = tk.Label(editWindow,text='数量：')
		lblCurrentPrice2 = tk.Label(editWindow,text=selectedValues[1])
		lblCurrentAmount = tk.Label(editWindow,text='价格：')
		lblCurrentAmount2 = tk.Label(editWindow,text=selectedValues[2])
		lblCurrentItem.place(x=0, y=0*int(height/13)+10, anchor='nw')
		lblCurrentPrice.place(x=0, y=1*int(height/13)+10, anchor='nw')
		lblCurrentAmount.place(x=240, y=1*int(height/13)+10, anchor='nw')
		lblCurrentItem2.place(x=100, y=0*int(height/13)+10, anchor='nw')
		lblCurrentPrice2.place(x=60, y=1*int(height/13)+10, anchor='nw')
		lblCurrentAmount2.place(x=300, y=1*int(height/13)+10, anchor='nw')
		lblCurrentInfo = tk.Label(editWindow,text='修改为：')
		lblCurrentItem = tk.Label(editWindow,text='项目：')
		cbbCurrentItem = ttk.Combobox(editWindow,width=48)
		lblCurrentPrice = tk.Label(editWindow,text='单价：')
		entCurrentPrice = tk.Entry(editWindow, width=int(480/32))
		lblCurrentAmount = tk.Label(editWindow,text='数量：')
		entCurrentAmount = tk.Entry(editWindow, width=int(480/32))
		lblCurrentInfo.place(x=0, y=3*int(height/13)+10, anchor='nw')
		lblCurrentItem.place(x=0, y=4*int(height/13)+10, anchor='nw')
		lblCurrentPrice.place(x=0, y=5*int(height/13)+10, anchor='nw')
		lblCurrentAmount.place(x=240, y=5*int(height/13)+10, anchor='nw')
		cbbCurrentItem.place(x=45, y=4*int(height/13)+10, anchor='nw')
		entCurrentPrice.place(x=45, y=5*int(height/13)+10, anchor='nw')
		entCurrentAmount.place(x=285, y=5*int(height/13)+10, anchor='nw')
		btnOK = tk.Button(editWindow, text="确定", width=int(width/50), height=1, command=setItem, compound="c")
		btnCancel = tk.Button(editWindow, text="取消", width=int(width/50), height=1, command=lambda:cancel(editWindow), compound="c")
		btnOK.place(x=160, y=6*int(height/13)+10, anchor='nw')
		btnCancel.place(x=320, y=6*int(height/13)+10, anchor='nw')
	else:
		messagebox.showerror(title='错误', message='请选择列表中的某一项进行修改！')

#main window initialize
mainWindow = tk.Tk()
mainWindow.geometry(str(width)+'x'+str(height))
mainWindow.resizable(False, False)
mainWindow.title("报价单生成工具"+' '+version)
mainWindow.iconbitmap("icon.ico")
mergeCheck = tk.IntVar()

#widget define
btnManage = tk.Button(mainWindow, text="管理类目物品", width=int(width/15), height=1, command=manageClick, compound="c")
btnImportExport = tk.Button(mainWindow, text="打开、保存和导出", width=int(width/15), height=1, command=importExportClick, compound="c")
treeview = ttk.Treeview(mainWindow, height=int(height/25), columns=("品名", "单价", "数量", "合计"))
treeview.column('#0', stretch=False, width=0, anchor='w')
treeview.column('品名', stretch=False, width=200, anchor='w')
treeview.column('单价', stretch=False, width=40, anchor='w')
treeview.column('数量', stretch=False, width=40, anchor='w')
treeview.column('合计', stretch=False, width=60, anchor='w')
treeview.heading("#0")
treeview.heading("品名", text="品名")
treeview.heading("单价", text="单价")
treeview.heading("数量", text="数量")
treeview.heading("合计", text="合计")
def handle_click(event):
    if treeview.identify_region(event.x, event.y) == "separator":
        return "break"
treeview.bind('<Button-1>', handle_click)
lblTitle = tk.Label(text='表头：')
lblAddress = tk.Label(text='地址：')
lblName = tk.Label(text='姓名：')
lblPhone = tk.Label(text='电话：')
lblDate = tk.Label(text='日期：')
lblContent = tk.Label(text='备注：')
entTitle = tk.Entry(mainWindow, width=int(width/18))
entAddress = tk.Entry(mainWindow, width=int(width/18))
entName = tk.Entry(mainWindow, width=int(width/18))
entPhone = tk.Entry(mainWindow, width=int(width/18))
defaultDate = tk.StringVar()
defaultDate.set(date)
entDate = tk.Entry(mainWindow, width=int(width/18), textvariable=defaultDate)
txtContent = tk.Text(mainWindow, width=int(width/18), height=3)
lblInsert = tk.Label(text='向表格中添加项目：')
lblItem = tk.Label(text='项目：')
lblAmount = tk.Label(text='单价：')
lblNum = tk.Label(text='数量：')
cbbItem = ttk.Combobox(width=int(width/18)-2)
entPrice = tk.Entry(mainWindow, width=int(width/32))
entAmount = tk.Entry(mainWindow, width=int(width/32))
ckbMerge = tk.Checkbutton(mainWindow, text="合并相同类目", variable=mergeCheck, onvalue=1, offvalue=0, command=merge, height=1)
btnInsert = tk.Button(mainWindow, text=">>插入>>", width=int(width/50), height=1, command=itemInsert, compound="c")
btnDelete = tk.Button(mainWindow, text="<<删除<<", width=int(width/50), height=1, command=itemDelete, compound="c")
btnEdit = tk.Button(mainWindow, text="<<修改>>", width=int(width/50), height=1, command=itemEdit, compound="c")
frm=tk.Frame(mainWindow,height=345,width=16)
#frm.grid(row=0,column=0,sticky='nsew')
bar = ttk.Scrollbar(frm, orient='vertical', command=treeview.yview)
treeview.configure(yscrollcommand=bar.set)
#bar.grid(row=1,column=2,rowspan=2,sticky='ns')

#main window
frm.place(x=width-5, y=height-11, anchor='se')
frm.pack_propagate(False)
bar.pack(side='right',fill='y',anchor='se')
btnImportExport.place(x=0, y=0, anchor='nw')
btnManage.place(x=width, y=0, anchor='ne')
treeview.place(x=width-22, y=height-10, anchor='se')
lblTitle.place(x=0, y=1*int(height/13)+10, anchor='nw')
lblAddress.place(x=0, y=2*int(height/13)+10, anchor='nw')
lblName.place(x=0, y=3*int(height/13)+10, anchor='nw')
lblPhone.place(x=0, y=4*int(height/13)+10, anchor='nw')
lblDate.place(x=0, y=5*int(height/13)+10, anchor='nw')
lblContent.place(x=0, y=6*int(height/13)+10, anchor='nw')
entTitle.place(x=45, y=1*int(height/13)+10, anchor='nw')
entAddress.place(x=45, y=2*int(height/13)+10, anchor='nw')
entName.place(x=45, y=3*int(height/13)+10, anchor='nw')
entPhone.place(x=45, y=4*int(height/13)+10, anchor='nw')
entDate.place(x=45, y=5*int(height/13)+10, anchor='nw')
txtContent.place(x=45, y=6*int(height/13)+10, anchor='nw')
lblInsert.place(x=0, y=8*int(height/13)+10, anchor='nw')
lblItem.place(x=0, y=9*int(height/13)+10, anchor='nw')
lblAmount.place(x=0, y=10*int(height/13)+10, anchor='nw')
lblNum.place(x=0, y=11*int(height/13)+10, anchor='nw')
cbbItem.place(x=45, y=9*int(height/13)+10, anchor='nw')
entPrice.place(x=45, y=10*int(height/13)+10, anchor='nw')
entAmount.place(x=45, y=11*int(height/13)+10, anchor='nw')
ckbMerge.place(x=0, y=12*int(height/13)+10, anchor='nw')
btnInsert.place(x=220, y=10*int(height/13)+10, anchor='nw')
btnEdit.place(x=220, y=11*int(height/13)+10, anchor='nw')
btnDelete.place(x=220, y=12*int(height/13)+10, anchor='nw')

#console test
print("version:"+version)
print("name:"+name)
print("address:"+address)
print("phone:"+phone)
print("date:"+date)

#default
mainWindow.mainloop()